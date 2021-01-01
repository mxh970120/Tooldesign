# -*- coding: utf-8 -*-

'''
Author: mxh970120
Date: 2020.12.31
'''

import sys
from PyQt5.QtWidgets import QWidget, QCheckBox, QApplication, QGridLayout, QPushButton, QMessageBox
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QIcon
import numpy as np
import openpyxl
import datetime


class Example(QWidget):

    def __init__(self):
        super().__init__()

        self.attributes = np.array([["学习", "没学习", "<1h", "1h~2h", ">2h"],
                              ["走路", "没走路", "<1万步", "1万步~10公里", ">10公里"],
                              ["编程", "没编程", "<1h", "1h~2h", ">2h"],
                              ["德语", "看了", "没看"],
                              ["英语", "看了", "没看"]])

        self.all_button_checkbox = np.copy(self.attributes)
        self.table = [0]*len(self.attributes)

        self.wb = openpyxl.load_workbook('daten.xlsx')
        now = datetime.datetime.now()
        if str(now.month) not in self.wb.sheetnames:
            self.wb.create_sheet(title=str(now.month))
        self.sheet = self.wb[str(now.month)]
        self.sheet.cell(row=1, column=1).value = '日期'
        self.sheet.cell(row=1, column=2).value = '时间'
        for i in range(len(self.attributes)):
            self.sheet.cell(row=1, column=i+3).value = self.attributes[i][0]

        self.initUI()

    def initUI(self):

        # QGridLayout的实例被创建并设置应用程序窗口的布局。
        grid = QGridLayout()
        self.setLayout(grid)
        # 设定属性
        for i in range(len(self.attributes)):
            for j, value in enumerate(self.attributes[i]):
                if j == 0:
                    button = QPushButton(value)
                    grid.addWidget(button, i, j)
                    self.all_button_checkbox[i][j] = button
                else:
                    cb = QCheckBox(value, self)
                    cb.stateChanged.connect(self.changeTitle)
                    grid.addWidget(cb, i, j)
                    self.all_button_checkbox[i][j] = cb

        verifyBtn = QPushButton("确定")
        verifyBtn.clicked.connect(self.verity)
        grid.addWidget(verifyBtn, len(self.attributes), 5)

        # 窗口的设置
        # 设置窗口的位置和大小(x, y, 长, 宽)
        self.setGeometry(300, 300, 300, 300)
        # 设置窗口的标题
        self.setWindowTitle('简易日志')
        # 设置窗口的图标，引用当前目录下的web.png图片
        self.setWindowIcon(QIcon('Icon.JPG'))
        # 显示窗口
        self.show()

    def changeTitle(self, state):
        sender = self.sender()
        position = [0, 0]
        for i in range(len(self.all_button_checkbox)):
            for j in self.all_button_checkbox[i]:
                if j == sender:
                    position = [i, j]
                    break

        if state == Qt.Checked:
            for checkbox in self.all_button_checkbox[position[0]]:
                if sender == checkbox:
                    self.table[position[0]] = sender.text()
                    print(self.table)
                else:
                    checkbox.setChecked(False)

    def verity(self, state):
        flag = True
        for item in self.table:
            if item == 0:
                flag = False
        if flag:
            message = "确定今天的记录了吗？"
        else:
            message = "记录尚未填写完整"

        reply = QMessageBox.question(self, 'Message', message,
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

        if reply == QMessageBox.Yes:
            self.speichern_table()

    def speichern_table(self):
        now = datetime.datetime.now()
        date = u'%s:%s:%s' % (now.year, now.month, now.day)
        nownow = u'%s:%s:%s' % (now.hour, now.minute, now.second)
        daten = [date, nownow] + self.table
        row = self.sheet.max_row
        cell = self.sheet.cell(row=row, column=1)
        if cell.value == date:
            self.sheet.delete_rows(row)
        self.sheet.append(daten)
        self.wb.save(filename="daten.xlsx")
        print("gespeichert")


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = Example()
    sys.exit(app.exec_())
